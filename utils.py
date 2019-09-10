import os
import pickle
import config as cfg
import numpy as np
import math
import openpyxl
import tqdm
from openpyxl.styles import PatternFill, colors, Font


def f1_score_max(gt, pred, thresh):
    from sklearn.metrics import precision_score, recall_score
    P = []
    R = []
    for i in thresh:
        new_pred = ((pred >= i) * 1).flatten()
        P.append(precision_score(gt.flatten(), new_pred))
        R.append(recall_score(gt.flatten(), new_pred))
    P = np.array(P).flatten()
    R = np.array(R).flatten()
    F1 = 2 * P * R / (P + R)
    F1_MAX = max(F1)
    if F1_MAX < 0 or math.isnan(F1_MAX):
        F1_MAX = 0
        F1_THRESH = 0
    else:
        idx_thresh = np.argmax(F1)
        F1_THRESH = thresh[idx_thresh]

    return F1, F1_MAX, F1_THRESH


def f1_score(gt, pred, F1_Thresh=0.5, files=None, median=False):
    from sklearn.metrics import f1_score as f1s
    if isinstance(gt, list):
        gt = np.array(gt)
    if isinstance(pred, list):
        pred = np.array(pred)
    # F1_Thresh = 0.5
    output = (pred > F1_Thresh) * 1.0
    F1 = f1s(gt, output)
    F1_MAX = F1

    if median:
        output_median3, output_median5, output_median7 = get_median(
            output, files)
        F1_median3 = f1s(gt, output_median3)
        F1_median5 = f1s(gt, output_median5)
        F1_median7 = f1s(gt, output_median7)

        return [F1], F1_MAX, F1_Thresh, F1_median3, F1_median5, F1_median7
    else:
        return [F1], F1_MAX, F1_Thresh


def get_median(pred, files, mode='all'):
    import pandas
    if mode == 'video':
        video_files, video_hist = get_unique_files(files)
        output_median = []
        output_median3 = []
        output_median5 = []
        output_median7 = []
        count = 0
        video = ''
        for i in range(len(video_hist)):
            assert video != video_files[count]
            video = video_files[count]
            range_frames = video_hist[video]
            output_median = pred[count:count + range_frames]
            count += range_frames

            output_median3.extend(
                np.array(
                    pandas.Series(output_median).rolling(
                        window=3, center=True).median().bfill().ffill()))
            output_median5.extend(
                np.array(
                    pandas.Series(output_median).rolling(
                        window=5, center=True).median().bfill().ffill()))
            output_median7.extend(
                np.array(
                    pandas.Series(output_median).rolling(
                        window=7, center=True).median().bfill().ffill()))
    elif mode == 'all':
        output_median3 = np.array(
            pandas.Series(pred).rolling(window=3,
                                        center=True).median().bfill().ffill())
        output_median5 = np.array(
            pandas.Series(pred).rolling(window=5,
                                        center=True).median().bfill().ffill())
        output_median7 = np.array(
            pandas.Series(pred).rolling(window=7,
                                        center=True).median().bfill().ffill())

    return output_median3, output_median5, output_median7


def get_unique_files(files):
    video_files = [os.path.dirname(line) for line in files]
    hist_video = {}
    for file_ in video_files:
        if file_ not in hist_video.keys():
            hist_video[file_] = 1
        else:
            hist_video[file_] += 1
    return video_files, hist_video


def whereAU(au):
    return np.where(np.array(cfg.AUs) == au)[0][0]


def createxls(config, mode):
    sheet = 'OF_' + config.OF_option if not config.HYDRA else 'Hydra_OF_'\
        + config.OF_option
    try:
        wb = openpyxl.load_workbook(
            config.xlsfile.replace('.xlsx', '_' + mode + '.xlsx'))
    except BaseException:
        wb = openpyxl.Workbook()
        wb.remove_sheet(wb.active)

    try:
        ws = wb.get_sheet_by_name(sheet)
    except BaseException:
        ws = wb.create_sheet(sheet)

    count = 1
    start_pos = [count]
    count = createSectionxls(ws, '0.5', count)
    start_pos.append(count)
    count = createSectionxls(ws, 'median3', count)
    start_pos.append(count)
    count = createSectionxls(ws, 'median5', count)
    start_pos.append(count)
    count = createSectionxls(ws, 'median7', count)
    start_pos.append(count)
    count = createSectionxls(ws, '1', count)
    return wb, ws, start_pos


def createSectionxls(worksheet, out='0.5', count=1):
    cell = cell2bold(worksheet['A' + str(count)])
    cell.value = '!' + out
    for i in range(3):
        cell = cell2bold(worksheet[chr(66 + i) + str(count)])
        cell.value = 'fold ' + str(i)
    cell = cell2Fcolor(worksheet['E' + str(count)], color='green')
    cell.value = 'mean'
    cell = cell2Fcolor(worksheet['F' + str(count)], color='green')
    cell.value = 'std'

    count += 1
    count = createPartAU(worksheet, count)
    for i in range(4):
        cell = cell2Fcolor(
            cell2bold(worksheet[chr(66 + i) + str(count)]), color='yellow')
        cell.value = '=AVERAGE({0}{1}:{0}{2})'.format(
            chr(66 + i), count - 12, count - 1)
    count += 3
    return count


def createPartAU(worksheet, count):
    for au in cfg.AUs:
        strAU = 'AU' + str(au).zfill(2)
        cell = cell2bold(worksheet['A' + str(count)])
        cell.value = strAU

        cell = worksheet['E' + str(count)]
        cell.value = '=AVERAGE(B{0}:D{0})'.format(count)

        cell = worksheet['F' + str(count)]
        cell.value = '=STDEV(B{0}:D{0})'.format(count)

        count += 1

    cell = worksheet['F' + str(count)]
    cell.value = '=STDEV(B{0}:D{0})'.format(count)

    cell = cell2color(worksheet['A' + str(count)], color='red')
    cell.value = 'MEAN'

    return count


def fillxls(config, worksheet, f1, count=1):
    au = whereAU(int(config.AU)) + 1 + count
    fold = int(config.fold) + 1
    cell = cell2color(worksheet[chr(65 + fold) + str(au)], color='blue')
    cell.value = f1


def cell2bold(cell):
    cell.font = Font(b=True)
    return cell


def cell2Fcolor(cell, color='red'):
    fill = PatternFill("solid", fgColor=getattr(colors, color.upper()))
    cell.fill = fill
    return cell


def cell2color(cell, color='red'):
    cell.font = Font(color=getattr(colors, color.upper()))
    return cell


def F1_TEST(config,
            data_loader,
            mode='TEST',
            thresh=0.5,
            OF=None,
            verbose=True):
    import torch.nn.functional as F
    PREDICTION = []
    GROUNDTRUTH = []
    FILES = []
    loss = []
    if verbose:
        print("-> xls results at " +
              config.xlsfile.replace('.xlsx', '_' + mode + '.xlsx'))
        workbook, worksheet, start_pos = createxls(config, mode)

    if OF is not None:
        of_loader = iter(OF)
    if verbose:
        print('\n================================')
        iterator = tqdm.tqdm(
            enumerate(data_loader),
            total=len(data_loader),
            ncols=10,
            desc="{} forward".format(mode.upper()))
    else:
        iterator = enumerate(data_loader)

    for i, (real_x, org_c, files) in iterator:
        if verbose and config.TEST_TXT:
            try:
                PREDICTION, GROUNDTRUTH, FILES = pickle.load(
                    open(config.pkl_data.format(mode.lower())))
                break
            except BaseException:
                PREDICTION, GROUNDTRUTH = pickle.load(
                    open(config.pkl_data.format(mode.lower())))
                break

        real_x = config.to_var(real_x, volatile=True)
        labels = org_c

        if OF is not None:
            of_x, of_c, of_files = next(of_loader)
            of_x = config.to_var(of_x, volatile=True)
            out_temp = config.C(real_x, OF=of_x)
        else:
            out_temp = config.C(real_x)

        output = F.sigmoid(out_temp)
        loss.append(config.LOSS(out_temp, config.to_var(org_c)))

        PREDICTION.extend(output.data.cpu().numpy().flatten().tolist())
        GROUNDTRUTH.extend(labels.cpu().numpy().astype(np.uint8).tolist())
        FILES.extend(['/'.join(file_.split('/')[-3:]) for file_ in files])

    if verbose:
        pickle.dump([PREDICTION, GROUNDTRUTH, FILES],
                    open(config.pkl_data.format(mode.lower()), 'w'))
    if verbose:
        print("")
        print >> config.f, ""

    PREDICTION = np.array(PREDICTION).flatten()
    GROUNDTRUTH = np.array(GROUNDTRUTH).flatten()

    prediction = PREDICTION
    groundtruth = GROUNDTRUTH
    if mode == 'TEST':
        _, F1_real5, F1_Thresh5, F1_median3, F1_median5, F1_median7 = \
            f1_score(
                groundtruth, prediction, 0.5, files=FILES, median=True)
        _, F1_real, F1_Thresh, F1_median3_th, F1_median5_th, F1_median7_th = \
            f1_score(
                np.array(groundtruth),
                np.array(prediction),
                thresh,
                files=FILES,
                median=True)
    else:
        _, F1_real, F1_Thresh, F1_median3, F1_median5, F1_median7 = \
            f1_score(
                np.array(groundtruth),
                np.array(prediction),
                thresh,
                files=FILES,
                median=True)
        F1_real5 = F1_real
        F1_median3_th = F1_median3
        F1_median5_th = F1_median5
        F1_median7_th = F1_median7
    _, F1_0, F1_Thresh_0 = f1_score(
        np.array(groundtruth), np.zeros_like(prediction), thresh)
    _, F1_1, F1_Thresh_1 = f1_score(
        np.array(groundtruth), np.ones_like(prediction), thresh)
    _, F1_MAX, F1_Thresh_max = f1_score_max(
        np.array(groundtruth), np.array(prediction), config.thresh)

    if verbose:
        fillxls(config, worksheet, F1_real5, start_pos[0])
        fillxls(config, worksheet, F1_median3, start_pos[1])
        fillxls(config, worksheet, F1_median5, start_pos[2])
        fillxls(config, worksheet, F1_median7, start_pos[3])
        fillxls(config, worksheet, F1_1, start_pos[4])
        workbook.close()

    string = "---> [%s - 0] AU%s F1: %.4f, Threshold: %.4f <---" % (
        mode, str(config.AU).zfill(2), F1_0, F1_Thresh_0)
    if verbose:
        print(string)
        print >> config.f, string

    string = "---> [%s - 1] AU%s F1: %.4f, Threshold: %.4f <---" % (
        mode, str(config.AU).zfill(2), F1_1, F1_Thresh_1)
    if verbose:
        print(string)
        print >> config.f, string

    string = "\n########\n#######  Threshold 0.5 ########\n########"
    if verbose:
        print(string)
        print >> config.f, string

    if mode == 'TEST':
        string = "---> [%s] AU%s F1: %.4f, Threshold: %.4f <---" % (
            mode, str(config.AU).zfill(2), F1_real5, F1_Thresh5)
        if verbose:
            print(string)
            print >> config.f, string

        string = "---> [%s] AU%s F1_median3: %.4f, Threshold: %.4f <---" % (
            mode, str(config.AU).zfill(2), F1_median3, F1_Thresh5)
        if verbose:
            print(string)
            print >> config.f, string

        string = "---> [%s] AU%s F1_median5: %.4f, Threshold: %.4f <---" % (
            mode, str(config.AU).zfill(2), F1_median5, F1_Thresh5)
        if verbose:
            print(string)
            print >> config.f, string

        string = "---> [%s] AU%s F1_median7: %.4f, Threshold: %.4f <---" % (
            mode, str(config.AU).zfill(2), F1_median7, F1_Thresh5)
        if verbose:
            print(string)
            print >> config.f, string

        string = "\n########\n#######  Threshold VAL ########\n########"
        if verbose:
            print(string)
            print >> config.f, string

    string = "---> [%s] AU%s F1: %.4f, Threshold: %.4f <---" % (
        mode, str(config.AU).zfill(2), F1_real, F1_Thresh)
    if verbose:
        print(string)
        print >> config.f, string

    string = "---> [%s] AU%s F1_median3: %.4f, Threshold: %.4f <---" % (
        mode, str(config.AU).zfill(2), F1_median3_th, F1_Thresh)
    if verbose:
        print(string)
        print >> config.f, string

    string = "---> [%s] AU%s F1_median5: %.4f, Threshold: %.4f <---" % (
        mode, str(config.AU).zfill(2), F1_median5_th, F1_Thresh)
    if verbose:
        print(string)
        print >> config.f, string

    string = "---> [%s] AU%s F1_median7: %.4f, Threshold: %.4f <---" % (
        mode, str(config.AU).zfill(2), F1_median7_th, F1_Thresh)
    if verbose:
        print(string)
        print >> config.f, string

    string = "\n########\n#######  Threshold MAX ########\n########"
    if verbose:
        print(string)
        print >> config.f, string

    # REAL F1_MAX
    string = "---> [%s] AU%s F1_MAX: %.4f, Threshold: %.4f <---" % (
        mode, str(config.AU).zfill(2), F1_MAX, F1_Thresh_max)
    if verbose:
        print(string)
        print >> config.f, string
        workbook.save(config.xlsfile.replace('.xlsx', '_' + mode + '.xlsx'))
    if mode == 'VAL':
        return F1_real, F1_MAX, F1_Thresh_max, np.array(loss).mean(), F1_1
    else:
        if verbose:
            print('================================')
        return F1_real, F1_MAX, F1_Thresh_max


def pdf2png(filename):
    from wand.image import Image
    from wand.color import Color
    with Image(filename="{}.pdf".format(filename), resolution=500) as img:
        with Image(
                width=img.width, height=img.height,
                background=Color("white")) as bg:
            bg.composite(img, 0, 0)
            bg.save(filename="{}.png".format(filename))
    os.remove('{}.pdf'.format(filename))
